# This program takes a number (N) from the command line and displays
# a NxN multiplication table in an excel spreadsheet

def multigrid():

    # need sys to read from command line,
    # need openpyxl to work with excel
    # need os for excel file location check

    import sys, openpyxl, os
    from openpyxl.styles import Font, Alignment

    N = 1  # default value

    # get a number from the command line
    
    if len(sys.argv) == 2: # check if there is one single argument
        N = sys.argv[1]    # set N to the value given
    else:
        print("You have not provided a single number, so a default of \
                1 will be used")

    # if the value is not a number, continue to prompt for one

    while not(str(N).isnumeric()):  
        N = int( input("A single number is required for this to run. \n\n \
                        ... Try again: ") )        

    # check if the excel sheet has been created yet

    name = "MultiTable in Excel.xlsx"

    if os.path.exists(name):
        wb= openpyxl.load_workbook(name)
        print("excel file exists")
        wb.create_sheet("new")
        del wb["Multiplication Table"]
        sheet = wb["new"]
        sheet.title= "Multiplication Table"
        sheet["A1"] = "X"
        sheet["A1"].font= Font(size=14, color = "00008B")
        sheet["A1"].alignment = Alignment(horizontal= "center")
        #wb.save()
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title ="Multiplication Table"
        sheet["A1"] = "X" + str(N)
        wb.save(name)
        print("excel file created")

    sheet = wb.active

    n = int(N)
   
    for number in range(2, n+2):                
        sheet.cell(1,number).value = number-1
        sheet.cell(number, 1).value = number-1

        sheet.cell(1,number).font = Font(bold=True, size=12, color = "FF0000")
        sheet.cell(number,1).font = Font(bold=True, size=12, color = "FF0000")

        sheet.cell(1,number).alignment = Alignment(horizontal= "center")
        sheet.cell(number,1).alignment = Alignment(horizontal= "center")

                                
    for row in range(2, n+2):    
        for col in range(2, n+2):   
            sheet.cell(row,col).value = (row-1)*(col-1)
            sheet.cell(row,col).alignment = Alignment(horizontal= "center")

                                                      
    wb.save(name)


    print("done")


multigrid()
        
